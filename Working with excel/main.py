import openpyxl
from openpyxl.styles import Font, PatternFill

# load an existing workbook
wb = openpyxl.load_workbook('practise.xlsx')
ws = wb.active

# apply font and fill formatting
bold_font = Font(bold=True)
fill_color = PatternFill(start_color='FFFF00', fill_type='solid')

ws['A1'].font = bold_font
ws['B1'].font = bold_font
ws['A1'].fill = fill_color
ws['B1'].fill = fill_color

# save the formatted workbook
wb.save('practise.xlsx')